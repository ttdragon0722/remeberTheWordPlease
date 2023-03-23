from json import dumps
from openpyxl import Workbook,load_workbook

wb = Workbook()
ws = wb.active
wb = load_workbook(filename="vocabulary.xlsx")
sheet = wb.worksheets[0]

result = {
    "vocabularies":[]
}

if __name__ == "__main__":


    for row in range(2,sheet.max_row+1):
        dictionary = {
            "vocabulary":sheet.cell(row,1).value,
            "partOfSpeech":sheet.cell(row,2).value,
            "chinese":sheet.cell(row,3).value
        }
        result["vocabularies"].append(dictionary)

    wb.close()  
    
    with open("vocabulary.json",'w',encoding="utf-8") as file :
        file.write(dumps(result,ensure_ascii=False))